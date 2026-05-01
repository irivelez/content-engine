#!/usr/bin/env python3
"""
Performance Loop — weekly metrics extractor for the content engine.

Reads Irina's LinkedIn .xlsx export (Creator Mode → Post Analytics → Export)
and the Beehiiv API (if credentials present). Writes a single JSON to
data/weekly-review/{YYYY-WW}-raw.json with activity IDs, metrics, and audience.

Does NOT cache LinkedIn post bodies. When voice/hook analysis needs a body,
agents fetch the LinkedIn URL live using the activity_id from this JSON.

Run:   python3 scripts/performance-loop.py [path/to/Content_*.xlsx]
"""

from __future__ import annotations

import json
import os
import sys
import re
import urllib.request
import urllib.error
from datetime import datetime, date, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("missing dep — run: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
ANALYTICS_DIR = ROOT / "data" / "post-analytics"
OUTPUT_DIR = ROOT / "data" / "weekly-review"
ENV_FILE = ROOT / ".env"

ACTIVITY_RE = re.compile(r"urn:li:activity:(\d+)")


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    if not ENV_FILE.exists():
        return env
    for line in ENV_FILE.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def latest_xlsx() -> Path:
    candidates = sorted(ANALYTICS_DIR.glob("Content_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        sys.exit(f"no xlsx in {ANALYTICS_DIR}")
    return candidates[0]


def parse_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%-m/%-d/%Y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def parse_period(xlsx_path: Path) -> tuple[str | None, str | None, str | None]:
    """From filename Content_YYYY-MM-DD_YYYY-MM-DD_*.xlsx → (start, end, iso_week)."""
    m = re.search(r"Content_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})", xlsx_path.name)
    if not m:
        return None, None, None
    start, end = m.group(1), m.group(2)
    iso_year, iso_week, _ = date.fromisoformat(end).isocalendar()
    return start, end, f"{iso_year}-W{iso_week:02d}"


def extract_activity_id(url: str) -> str | None:
    if not isinstance(url, str):
        return None
    m = ACTIVITY_RE.search(url)
    return m.group(1) if m else None


def read_linkedin(xlsx_path: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)

    discovery: dict[str, Any] = {}
    ws = wb["DISCOVERY"]
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        key, val = row[0], row[1] if len(row) > 1 else None
        if "Performance" in str(key) and val:
            discovery["period_label"] = str(val)
        elif key == "Impressions":
            discovery["impressions"] = int(val) if val else None
        elif key == "Members reached":
            discovery["members_reached"] = int(val) if val else None

    engagement: list[dict[str, Any]] = []
    ws = wb["ENGAGEMENT"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        d = parse_date(row[0])
        if not d:
            continue
        engagement.append({
            "date": d,
            "impressions": int(row[1]) if row[1] is not None else 0,
            "engagements": int(row[2]) if row[2] is not None else 0,
        })

    by_engagements: list[dict[str, Any]] = []
    by_impressions: list[dict[str, Any]] = []
    ws = wb["TOP POSTS"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        if row[0]:
            aid = extract_activity_id(row[0])
            by_engagements.append({
                "activity_id": aid,
                "url": row[0],
                "publish_date": parse_date(row[1]),
                "engagements": int(row[2]) if row[2] is not None else 0,
            })
        if len(row) > 4 and row[4]:
            aid = extract_activity_id(row[4])
            by_impressions.append({
                "activity_id": aid,
                "url": row[4],
                "publish_date": parse_date(row[5]),
                "impressions": int(row[6]) if row[6] is not None else 0,
            })

    followers_total: int | None = None
    new_per_day: list[dict[str, Any]] = []
    ws = wb["FOLLOWERS"]
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        if "Total followers" in str(row[0]) and row[1] is not None:
            followers_total = int(row[1])
            continue
        d = parse_date(row[0])
        if d and row[1] is not None:
            new_per_day.append({"date": d, "new_followers": int(row[1])})

    def to_pct(v: Any) -> float | None:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if "<" in s:
            return 0.005  # "< 1%" → midpoint approximation
        try:
            return float(s.rstrip("%")) / 100.0 if "%" in s else float(s)
        except ValueError:
            return None

    demographics: dict[str, list[dict[str, Any]]] = {}
    ws = wb["DEMOGRAPHICS"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        cat, value, pct = row[0], row[1], row[2]
        if not cat or value is None:
            continue
        key = str(cat).lower().replace(" ", "_")
        demographics.setdefault(key, []).append({
            "value": str(value),
            "percentage": to_pct(pct),
        })

    return {
        "discovery": discovery,
        "engagement_per_day": engagement,
        "top_posts_by_engagements": by_engagements,
        "top_posts_by_impressions": by_impressions,
        "total_followers": followers_total,
        "new_followers_per_day": new_per_day,
        "demographics": demographics,
    }


def http_get_json(url: str, headers: dict[str, str], timeout: int = 20) -> Any:
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def read_beehiiv(env: dict[str, str]) -> dict[str, Any]:
    api_key = env.get("BEEHIIV_API_KEY", "")
    pub_id = env.get("BEEHIIV_PUBLICATION_ID", "")

    if not api_key or not pub_id:
        return {
            "status": "skipped",
            "reason": "BEEHIIV_API_KEY or BEEHIIV_PUBLICATION_ID missing in .env",
            "subscriptions": None,
            "posts": [],
        }

    base = "https://api.beehiiv.com/v2"
    headers = {"Authorization": f"Bearer {api_key}", "Accept": "application/json"}

    out: dict[str, Any] = {"status": "ok", "subscriptions": None, "posts": []}

    try:
        subs = http_get_json(f"{base}/publications/{pub_id}/subscriptions?limit=1", headers)
        out["subscriptions"] = {
            "total": (subs.get("total_results") or subs.get("total") or
                      (subs.get("page", {}) or {}).get("total_results")),
            "raw_keys": list(subs.keys()),
        }
    except (urllib.error.URLError, urllib.error.HTTPError, json.JSONDecodeError) as e:
        out["subscriptions"] = {"error": str(e)}

    try:
        posts = http_get_json(f"{base}/publications/{pub_id}/posts?limit=20&order_by=created", headers)
        items = posts.get("data", []) if isinstance(posts, dict) else []
        for p in items:
            out["posts"].append({
                "id": p.get("id"),
                "title": p.get("title"),
                "status": p.get("status"),
                "publish_date": p.get("publish_date"),
                "web_url": p.get("web_url"),
                "stats": p.get("stats", {}),
            })
    except (urllib.error.URLError, urllib.error.HTTPError, json.JSONDecodeError) as e:
        out["posts_error"] = str(e)

    return out


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else latest_xlsx()
    if not xlsx_path.exists():
        sys.exit(f"not found: {xlsx_path}")

    start, end, iso_week = parse_period(xlsx_path)
    env = load_env()

    linkedin = read_linkedin(xlsx_path)
    beehiiv = read_beehiiv(env)

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_xlsx": xlsx_path.name,
        "snapshot_period": {"start": start, "end": end, "iso_week": iso_week},
        "linkedin": linkedin,
        "beehiiv": beehiiv,
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"{iso_week or 'latest'}-raw.json"
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False))

    li = payload["linkedin"]
    print(f"wrote {out_path.relative_to(ROOT)}")
    print(f"  source     : {xlsx_path.name}")
    print(f"  iso_week   : {iso_week}")
    print(f"  followers  : {li.get('total_followers')}")
    print(f"  impressions: {li.get('discovery', {}).get('impressions')}")
    print(f"  top by imp : {len(li.get('top_posts_by_impressions', []))} posts")
    print(f"  top by eng : {len(li.get('top_posts_by_engagements', []))} posts")
    print(f"  beehiiv    : {beehiiv.get('status')}")


if __name__ == "__main__":
    main()
